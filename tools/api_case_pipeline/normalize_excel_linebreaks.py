"""在不重写工作簿结构的前提下规范化 XLSX 单元格换行。"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
CONFIG = ROOT / "tools" / "api_case_pipeline" / "config" / "storage_scope.json"
BACKUP = ROOT / "temp_scripts" / "存储管理单接口测试用例.before_linebreak_cleanup_20260805.xlsx"

CR_ENTITY_RUN = re.compile(r"(?:&#13;)+(?:&#10;)?")
X000D_RUN = re.compile(r"(?:_x000D_)+(?:\r\n|\r|\n)?")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def count_affected_cells(workbook_path: Path) -> tuple[int, dict[str, int]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    by_sheet: dict[str, int] = {}
    try:
        for sheet in workbook.worksheets:
            count = sum(
                1
                for row in sheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str)
                and ("\r" in cell.value or "_x000D_" in cell.value)
            )
            if count:
                by_sheet[sheet.title] = count
    finally:
        workbook.close()
    return sum(by_sheet.values()), by_sheet


def normalize_xml(data: bytes) -> tuple[bytes, int]:
    text = data.decode("utf-8")
    text, entity_changes = CR_ENTITY_RUN.subn("&#10;", text)
    text, escaped_changes = X000D_RUN.subn("&#10;", text)
    return text.encode("utf-8"), entity_changes + escaped_changes


def build_normalized_workbook(source: Path, destination: Path) -> dict[str, int]:
    changed_entries: dict[str, int] = {}
    with ZipFile(source, "r") as archive, ZipFile(destination, "w") as output:
        for info in archive.infolist():
            data = archive.read(info.filename)
            should_scan = (
                info.filename == "xl/sharedStrings.xml"
                or (
                    info.filename.startswith("xl/worksheets/")
                    and info.filename.endswith(".xml")
                )
            )
            if should_scan:
                data, changes = normalize_xml(data)
                if changes:
                    changed_entries[info.filename] = changes
            kwargs = {"compress_type": info.compress_type}
            if info.compress_type == ZIP_DEFLATED:
                kwargs["compresslevel"] = 9
            output.writestr(info, data, **kwargs)
    with ZipFile(destination, "r") as archive:
        bad_entry = archive.testzip()
        if bad_entry:
            raise RuntimeError(f"ZIP 校验失败: {bad_entry}")
    return changed_entries


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="实际写入；默认只检查")
    args = parser.parse_args()

    config = json.loads(CONFIG.read_text(encoding="utf-8"))
    workbook_path = ROOT / config["excel"]
    affected_before, sheets_before = count_affected_cells(workbook_path)
    result = {
        "excel": str(workbook_path),
        "affectedCellsBefore": affected_before,
        "sheetsBefore": sheets_before,
        "applied": False,
    }
    if not args.apply:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    if BACKUP.exists():
        if sha256(BACKUP) != sha256(workbook_path):
            raise RuntimeError(f"已有备份与当前文件不同，拒绝继续: {BACKUP}")
    else:
        shutil.copy2(workbook_path, BACKUP)

    temporary = workbook_path.with_suffix(".normalize.tmp.xlsx")
    try:
        changed_entries = build_normalized_workbook(workbook_path, temporary)
        if not changed_entries:
            temporary.unlink(missing_ok=True)
            result["backup"] = str(BACKUP)
            result["message"] = "没有需要规范化的 XML 换行"
            print(json.dumps(result, ensure_ascii=False, indent=2))
            return
        os.replace(temporary, workbook_path)
    finally:
        temporary.unlink(missing_ok=True)

    affected_after, sheets_after = count_affected_cells(workbook_path)
    if affected_after:
        raise RuntimeError(f"规范化后仍有 {affected_after} 个异常换行单元格")
    result.update(
        {
            "applied": True,
            "backup": str(BACKUP),
            "changedEntries": changed_entries,
            "affectedCellsAfter": affected_after,
            "sheetsAfter": sheets_after,
            "sha256Before": sha256(BACKUP),
            "sha256After": sha256(workbook_path),
            "sizeBefore": BACKUP.stat().st_size,
            "sizeAfter": workbook_path.stat().st_size,
        }
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
