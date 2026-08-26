#!/usr/bin/env python3
"""Convert PRD test case JSON files to Excel workbooks and merge them."""

import argparse
import json
import os
import re
import sys
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "编号",
    "模块",
    "用例名称",
    "优先级",
    "前置条件",
    "操作步骤",
    "输入数据",
    "预期结果",
    "验证结果",
    "备注",
    "bug",  # 手工填写列，生成时留空；已有内容保留
]
HEADER_FONT_SIZE = 12
HEADER_ROW_HEIGHT = 30
# 表头灰色填充：与 WPS/Excel 内置“白色，背景1，深色15%”一致（theme=0, tint=-0.15，约 #D9D9D9）
HEADER_FILL = PatternFill(
    patternType="solid",
    fgColor=Color(theme=0, tint=-0.15),
    bgColor=Color(rgb="00F0F0F0"),
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
# 验证结果列允许值（执行时可点选）
VERIFY_RESULT_OPTIONS = ["通过", "失败", "阻塞", "未执行"]
# 验证结果列按取值着色（字体颜色）：通过绿、失败红、阻塞黄、未执行灰
VERIFY_RESULT_COLORS = {
    "通过": "008000",
    "失败": "FF0000",
    "阻塞": "FFC000",
    "未执行": "808080",
}
# 每列列宽上限（Excel 宽度单位），防止长文本把表格撑得过宽
COLUMN_CAPS = {
    "编号": 20,
    "模块": 26,
    "用例名称": 50,
    "优先级": 10,
    "前置条件": 44,
    "操作步骤": 48,
    "输入数据": 46,
    "预期结果": 54,
    "验证结果": 12,
    "备注": 46,
    "bug": 46,
}


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item).strip() for item in value if str(item).strip())
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if "\n" in text:
        return text
    return re.sub(r"([；;])\s*(?=\d+\s*[.、])", r"\1\n", text)


def read_cases(path):
    with open(path, "r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise ValueError(f"{path} must contain a JSON array of cases")
    for index, item in enumerate(data):
        if "bug" not in item:
            item["bug"] = ""
        missing = [header for header in HEADERS if header not in item]
        if missing:
            raise ValueError(f"{path} case {index + 1}: missing fields {missing}")
    return data


def display_width(text):
    """Estimate visible width in Excel width units (CJK characters count as 2)."""
    return sum(
        2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
        for ch in str(text)
    )


def style_header(ws):
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for cell in ws[1]:
        cell.font = Font(bold=True, size=HEADER_FONT_SIZE)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center")


def write_workbook(path, cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(HEADERS)
    style_header(ws)
    for case in cases:
        ws.append([cell_text(case[header]) for header in HEADERS])
    apply_layout(ws)
    wb.save(path)
    return len(cases)


def apply_layout(ws):
    for index, header in enumerate(HEADERS, start=1):
        letter = get_column_letter(index)
        width = 8
        for cell in ws[letter]:
            for line in str(cell.value or "").split("\n"):
                width = max(width, display_width(line) + 2)
        ws.column_dimensions[letter].width = min(width, COLUMN_CAPS[header])
    # 全部数据单元格（含留空的 bug 列）统一细框线
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            horizontal = "center" if cell.column in (4, 9) else "left"
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal=horizontal
            )
    add_verify_validation(ws)
    add_verify_conditional_format(ws)
    ws.freeze_panes = "A2"




def add_verify_validation(ws):
    """给 验证结果 列（I 列）添加下拉，允许 通过/失败/阻塞/未执行。"""
    if ws.max_row < 2:
        return
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(VERIFY_RESULT_OPTIONS) + '"',
        allow_blank=True,
    )
    dv.add(f"I2:I{ws.max_row}")
    ws.add_data_validation(dv)




def add_verify_conditional_format(ws):
    """验证结果列按取值着色：通过绿、失败红、阻塞黄（琥珀）、未执行灰。"""
    if ws.max_row < 2:
        return
    rng = f"I2:I{ws.max_row}"
    for value, color in VERIFY_RESULT_COLORS.items():
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{value}"'], font=Font(color=color)),
        )


def merge_workbooks(output_path, module_paths):
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(HEADERS)
    style_header(ws)
    total = 0
    for path in module_paths:
        source = load_workbook(path, read_only=True, data_only=True)
        sheet = source.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ws.append(list(row))
        total += max(0, sheet.max_row - 1)
        source.close()
    apply_layout(ws)
    wb.save(output_path)
    return total


def main():
    parser = argparse.ArgumentParser(description="Build test case Excel files from JSON")
    parser.add_argument("--project", required=True, help="project folder name")
    parser.add_argument("--cases-dir", required=True, help="directory containing case JSON files")
    parser.add_argument("--output-dir", required=True, help="directory for Excel output")
    parser.add_argument("--merge-name", default="测试用例.xlsx", help="merged workbook name")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    json_files = sorted(
        name for name in os.listdir(args.cases_dir) if name.lower().endswith(".json")
    )
    if not json_files:
        print(f"no JSON files found in {args.cases_dir}")
        sys.exit(1)

    module_paths = []
    total = 0
    for name in json_files:
        cases = read_cases(os.path.join(args.cases_dir, name))
        output_path = os.path.join(args.output_dir, os.path.splitext(name)[0] + ".xlsx")
        count = write_workbook(output_path, cases)
        module_paths.append(output_path)
        total += count
        print(f"{output_path}: {count} cases")

    merged_path = os.path.join(args.output_dir, args.merge_name)
    merged_total = merge_workbooks(merged_path, module_paths)
    print(f"{merged_path}: {merged_total} cases")
    print(f"project={args.project} total={total}")


if __name__ == "__main__":
    main()
