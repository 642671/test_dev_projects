"""
文件处理工具
支持 YAML 和 Excel 文件的读写操作
"""
import os
import yaml
from openpyxl import Workbook, load_workbook
from common.logger import get_logger

logger = get_logger("FileHandler")


class YAMLHandler:
    """YAML 文件处理"""

    @staticmethod
    def read(file_path):
        """
        读取 YAML 文件
        :param file_path: YAML 文件路径
        :return: 解析后的数据（字典或列表），读取失败返回 None
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"YAML 文件不存在: {file_path}")
                return None
            with open(file_path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
            logger.info(f"成功读取 YAML 文件: {file_path}")
            return data
        except yaml.YAMLError as e:
            logger.error(f"YAML 文件解析失败: {file_path}, 错误: {e}")
            return None
        except Exception as e:
            logger.error(f"读取 YAML 文件异常: {file_path}, 错误: {e}")
            return None

    @staticmethod
    def write(file_path, data):
        """
        写入 YAML 文件
        :param file_path: 输出文件路径
        :param data: 要写入的数据（字典或列表）
        """
        try:
            # 确保目标目录存在
            dir_path = os.path.dirname(file_path)
            if dir_path:
                os.makedirs(dir_path, exist_ok=True)
            with open(file_path, "w", encoding="utf-8") as f:
                yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
            logger.info(f"成功写入 YAML 文件: {file_path}")
        except Exception as e:
            logger.error(f"写入 YAML 文件失败: {file_path}, 错误: {e}")
            raise

    @staticmethod
    def read_all(file_path):
        """
        读取多文档 YAML（包含多个 --- 分隔的文档）
        :param file_path: YAML 文件路径
        :return: 文档列表，每个元素为一个文档的解析结果
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"YAML 文件不存在: {file_path}")
                return None
            with open(file_path, "r", encoding="utf-8") as f:
                documents = list(yaml.safe_load_all(f))
            logger.info(f"成功读取多文档 YAML 文件: {file_path}, 共 {len(documents)} 个文档")
            return documents
        except yaml.YAMLError as e:
            logger.error(f"多文档 YAML 解析失败: {file_path}, 错误: {e}")
            return None
        except Exception as e:
            logger.error(f"读取多文档 YAML 异常: {file_path}, 错误: {e}")
            return None


class ExcelHandler:
    """Excel 文件处理（基于 openpyxl）"""

    @staticmethod
    def read(file_path, sheet_name=None):
        """
        读取 Excel 文件，返回列表（每行为一个字典）
        第一行作为表头
        :param file_path: Excel 文件路径
        :param sheet_name: 工作表名称，默认读取活动工作表
        :return: 字典列表，每个字典对应一行数据；读取失败返回 None
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"Excel 文件不存在: {file_path}")
                return None
            wb = load_workbook(file_path, read_only=True)
            # 选择工作表
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    logger.error(f"工作表 '{sheet_name}' 不存在于文件: {file_path}")
                    wb.close()
                    return None
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 读取所有行
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                logger.warning(f"Excel 文件为空: {file_path}")
                return []

            # 第一行作为表头
            headers = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                data.append(row_dict)

            logger.info(f"成功读取 Excel 文件: {file_path}, 共 {len(data)} 行数据")
            return data
        except Exception as e:
            logger.error(f"读取 Excel 文件异常: {file_path}, 错误: {e}")
            return None

    @staticmethod
    def write(file_path, data, sheet_name="Sheet1", headers=None):
        """
        写入 Excel 文件
        :param file_path: 文件路径
        :param data: 数据列表（每项为字典或列表）
        :param sheet_name: 工作表名称
        :param headers: 表头列表（如果 data 为字典列表可自动推断）
        """
        try:
            # 确保目标目录存在
            dir_path = os.path.dirname(file_path)
            if dir_path:
                os.makedirs(dir_path, exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # 确定表头
            if headers is None and data and isinstance(data[0], dict):
                headers = list(data[0].keys())

            # 写入表头
            if headers:
                ws.append(headers)

            # 写入数据行
            for item in data:
                if isinstance(item, dict):
                    # 按表头顺序提取值
                    row = [item.get(h, "") for h in headers] if headers else list(item.values())
                elif isinstance(item, (list, tuple)):
                    row = list(item)
                else:
                    row = [item]
                ws.append(row)

            wb.save(file_path)
            logger.info(f"成功写入 Excel 文件: {file_path}, 共 {len(data)} 行数据")
        except Exception as e:
            logger.error(f"写入 Excel 文件失败: {file_path}, 错误: {e}")
            raise

    @staticmethod
    def append_row(file_path, row_data, sheet_name=None):
        """
        向已有 Excel 追加一行数据
        :param file_path: Excel 文件路径
        :param row_data: 行数据（列表或字典）
        :param sheet_name: 工作表名称，默认使用活动工作表
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"Excel 文件不存在，无法追加: {file_path}")
                raise FileNotFoundError(f"文件不存在: {file_path}")

            wb = load_workbook(file_path)
            # 选择工作表
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    logger.error(f"工作表 '{sheet_name}' 不存在于文件: {file_path}")
                    wb.close()
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 如果是字典类型，按表头顺序排列
            if isinstance(row_data, dict):
                # 获取表头（第一行）
                headers = [cell.value for cell in ws[1]]
                row = [row_data.get(h, "") for h in headers]
            elif isinstance(row_data, (list, tuple)):
                row = list(row_data)
            else:
                row = [row_data]

            ws.append(row)
            wb.save(file_path)
            logger.info(f"成功向 Excel 文件追加一行数据: {file_path}")
        except (FileNotFoundError, ValueError):
            raise
        except Exception as e:
            logger.error(f"向 Excel 追加数据失败: {file_path}, 错误: {e}")
            raise
